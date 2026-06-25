import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
import os

excel_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "QA_Test_Suite.xlsx")

def update_master_report():
    if not os.path.exists(excel_file):
        print("Excel file not found.")
        return

    sheets = ["Selenium_300", "Appium_300", "API_300", "Validation_300", "Performance_300"]
    
    total_passed = 0
    total_failed = 0
    total_blocked = 0
    total_skipped = 0
    total_duration_s = 0.0
    
    try:
        for sheet_name in sheets:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Count statuses
            if 'Status' in df.columns:
                counts = df['Status'].value_counts()
                total_passed += counts.get('Pass', 0)
                total_failed += counts.get('Fail', 0)
                total_blocked += counts.get('Blocked', 0)
                total_skipped += counts.get('Skipped', 0)
            
            # Calculate duration
            if 'Execution Time' in df.columns:
                # Execution Time is expected to be string like '1.25s'
                def parse_duration(val):
                    if pd.isna(val): return 0.0
                    if isinstance(val, str) and val.endswith('s'):
                        try:
                            return float(val[:-1])
                        except ValueError:
                            return 0.0
                    return 0.0
                
                total_duration_s += df['Execution Time'].apply(parse_duration).sum()
                
    except Exception as e:
        print(f"Error reading sheets: {e}")
        return

    total_executed = total_passed + total_failed + total_blocked + total_skipped
    pass_pct = (total_passed / 1500) * 100 if total_executed > 0 else 0
    
    # Write to master sheet
    wb = openpyxl.load_workbook(excel_file)
    master_sheet = wb["Master_Report"]
    
    master_sheet.cell(row=2, column=2).value = total_passed
    master_sheet.cell(row=3, column=2).value = total_failed
    master_sheet.cell(row=4, column=2).value = total_skipped
    master_sheet.cell(row=5, column=2).value = total_blocked
    master_sheet.cell(row=6, column=2).value = f"{pass_pct:.2f}%"
    master_sheet.cell(row=7, column=2).value = f"{total_duration_s:.2f}s"
    
    # Add a simple Pie Chart
    chart = PieChart()
    chart.title = "Test Execution Results"
    labels = Reference(master_sheet, min_col=1, min_row=2, max_row=5)
    data = Reference(master_sheet, min_col=2, min_row=1, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    
    master_sheet.add_chart(chart, "D2")
    
    wb.save(excel_file)
    print("Master Report Successfully Updated with Execution Results and Charts.")

if __name__ == "__main__":
    update_master_report()
