import pytest
import os
from datetime import datetime
import openpyxl

excel_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "QA_Test_Suite.xlsx")
test_results_cache = []

def get_sheet_by_test_id(test_id):
    if test_id.startswith("SEL"): return "Selenium_300"
    if test_id.startswith("APP"): return "Appium_300"
    if test_id.startswith("API"): return "API_300"
    if test_id.startswith("VAL"): return "Validation_300"
    if test_id.startswith("PERF"): return "Performance_300"
    return None

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    
    if rep.when == "call":
        test_id = item.callspec.id if hasattr(item, "callspec") else item.name
        
        status = "Pass" if rep.passed else "Fail"
        actual_result = "Test executed successfully." if rep.passed else "Test failed during execution."
        error_msg = str(rep.longreprtext)[:500] if rep.failed else ""
        log_path = f"reports/logs/{test_id}.log" if rep.failed else ""
        
        duration = rep.duration
        
        test_results_cache.append({
            "test_id": test_id,
            "status": status,
            "actual_result": actual_result,
            "duration": duration,
            "error_msg": error_msg,
            "log_path": log_path,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

@pytest.hookimpl(tryfirst=True)
def pytest_sessionfinish(session, exitstatus):
    if not os.path.exists(excel_file):
        print(f"Excel file not found at {excel_file}")
        return

    try:
        workbook = openpyxl.load_workbook(excel_file)
        
        for result in test_results_cache:
            sheet_name = get_sheet_by_test_id(result["test_id"])
            if not sheet_name:
                continue
                
            sheet = workbook[sheet_name]
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == result["test_id"]:
                    sheet.cell(row=row, column=8).value = result["actual_result"]
                    sheet.cell(row=row, column=9).value = result["status"]
                    sheet.cell(row=row, column=10).value = f"{result['duration']:.2f}s"
                    sheet.cell(row=row, column=11).value = result["date"]
                    sheet.cell(row=row, column=16).value = result["log_path"]
                    sheet.cell(row=row, column=18).value = result["error_msg"]
                    break
                    
        workbook.save(excel_file)
        print(f"Successfully saved {len(test_results_cache)} test results to Excel.")
    except Exception as e:
        print(f"Failed to batch update Excel: {e}")
